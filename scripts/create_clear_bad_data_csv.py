"""
Create a CSV to clear the incorrectly imported recipe data from Salesforce
This will blank out Ingredients__c, Instructions__c, and Recipe_Content__c for the 63 meals
"""
from openpyxl import load_workbook
import csv

def create_clear_csv(user_excel, output_csv):
    """Create CSV to clear bad data from Salesforce"""

    # Load user's Excel file to get the IDs
    wb = load_workbook(user_excel)
    ws = wb.active

    clear_records = []

    for row_num in range(2, ws.max_row + 1):
        sf_id = ws.cell(row=row_num, column=1).value

        if sf_id and isinstance(sf_id, str) and sf_id.startswith('a1'):
            # Add record with blank values to clear the fields
            clear_records.append({
                'Id': sf_id,
                'Ingredients__c': '',
                'Instructions__c': '',
                'Recipe_Content__c': ''
            })

    # Write CSV
    if clear_records:
        with open(output_csv, 'w', newline='', encoding='utf-8') as f:
            fieldnames = ['Id', 'Ingredients__c', 'Instructions__c', 'Recipe_Content__c']
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(clear_records)

        print("="*80)
        print("CLEAR BAD DATA CSV CREATED")
        print("="*80)
        print(f"File: {output_csv}")
        print(f"Records: {len(clear_records)}")
        print()
        print("This CSV will CLEAR (blank out) the recipe data from these meals:")
        print("- Ingredients__c")
        print("- Instructions__c")
        print("- Recipe_Content__c")
        print()
        print("="*80)
        print("HOW TO USE THIS FILE:")
        print("="*80)
        print("1. Login to Workbench: https://workbench.developerforce.com")
        print("2. Data -> Update")
        print("3. Object: Meal__c")
        print("4. Upload this CSV")
        print("5. Map fields:")
        print("   - Id -> Record ID")
        print("   - Ingredients__c -> Ingredients")
        print("   - Instructions__c -> Instructions")
        print("   - Recipe_Content__c -> Recipe Content")
        print("6. Click Update")
        print()
        print("This will remove the incorrect data so you can re-import correctly.")
        print()

if __name__ == '__main__':
    user_excel = r'C:\Users\Abbyl\OneDrive\Desktop\Receips\new\existing_meals_export2.xlsx'
    output_csv = r'C:\Users\Abbyl\OneDrive\Desktop\Salesforce Training\Assistant\Assistant\data\CLEAR_BAD_DATA.csv'

    print("Creating CSV to clear bad recipe data from Salesforce...")
    print()

    create_clear_csv(user_excel, output_csv)

    print("IMPORTANT:")
    print("- Make sure you have a backup before using this!")
    print("- This will ERASE recipe data from 63 meals")
    print("- You can then re-import with the correct data")
    print()
