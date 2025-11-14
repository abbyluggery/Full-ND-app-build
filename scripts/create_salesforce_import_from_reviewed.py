"""
Create Salesforce import CSV from the reviewed Excel file
Matches recipe names to existing Salesforce meal IDs
"""
from openpyxl import load_workbook
import csv
import re

def create_salesforce_import_csv(reviewed_excel, existing_meals_csv, output_csv):
    """Create final Salesforce import CSV with IDs"""

    # Load existing Salesforce meals to get IDs
    salesforce_meals = {}
    with open(existing_meals_csv, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Store both exact name and normalized name
            name = row['Name'].strip()
            salesforce_meals[name.lower()] = {
                'Id': row['Id'],
                'Name': name
            }

    print(f"Loaded {len(salesforce_meals)} existing Salesforce meals")
    print()

    # Load reviewed Excel file
    wb = load_workbook(reviewed_excel)
    ws = wb.active

    # Get headers
    headers = [cell.value for cell in ws[1]]

    # Find column indices
    id_col = headers.index('Id') + 1 if 'Id' in headers else None
    suggested_name_col = headers.index('Suggested_Salesforce_Name') + 1
    ingredients_col = headers.index('Ingredients') + 1
    instructions_col = headers.index('Instructions') + 1
    recipe_content_col = headers.index('Recipe_Content') + 1

    # Process each row
    import_records = []
    matched_count = 0
    new_records_count = 0
    skipped_count = 0

    for row_num in range(2, ws.max_row + 1):
        # Get Salesforce ID (if provided by user)
        user_provided_id = ws.cell(row=row_num, column=id_col).value if id_col else None

        # Get suggested Salesforce name
        suggested_name = ws.cell(row=row_num, column=suggested_name_col).value

        # Skip if no suggested name
        if not suggested_name:
            skipped_count += 1
            continue

        suggested_name = str(suggested_name).strip()

        # Get recipe data
        ingredients = ws.cell(row=row_num, column=ingredients_col).value or ''
        instructions = ws.cell(row=row_num, column=instructions_col).value or ''
        recipe_content = ws.cell(row=row_num, column=recipe_content_col).value or ''

        # Determine Salesforce ID
        salesforce_id = None

        # First, check if user provided an ID
        if user_provided_id and isinstance(user_provided_id, str) and user_provided_id.startswith('a1'):
            salesforce_id = user_provided_id
            matched_count += 1
            print(f"USER PROVIDED ID: {suggested_name} -> {salesforce_id}")

        # Otherwise, try to match by name
        elif suggested_name.lower() in salesforce_meals:
            salesforce_id = salesforce_meals[suggested_name.lower()]['Id']
            matched_count += 1
            print(f"MATCHED: {suggested_name} -> {salesforce_id}")

        # If no match, check if it's marked as NEW
        elif '(NEW)' in suggested_name.upper() or 'NEW' in suggested_name.upper():
            new_records_count += 1
            print(f"NEW RECIPE (needs Salesforce record first): {suggested_name}")
            continue
        else:
            # Try fuzzy match
            found = False
            suggested_lower = suggested_name.lower()
            for sf_name_lower, sf_data in salesforce_meals.items():
                # Check if key words match
                if len(suggested_lower) > 5 and suggested_lower[:5] in sf_name_lower:
                    salesforce_id = sf_data['Id']
                    matched_count += 1
                    print(f"FUZZY MATCH: {suggested_name} -> {sf_data['Name']} ({salesforce_id})")
                    found = True
                    break

            if not found:
                new_records_count += 1
                print(f"NO MATCH (needs Salesforce record): {suggested_name}")
                continue

        # Add to import records
        if salesforce_id:
            import_records.append({
                'Id': salesforce_id,
                'Ingredients__c': str(ingredients).strip() if ingredients else '',
                'Instructions__c': str(instructions).strip() if instructions else '',
                'Recipe_Content__c': str(recipe_content).strip() if recipe_content else ''
            })

    # Write import CSV
    if import_records:
        with open(output_csv, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Id', 'Ingredients__c', 'Instructions__c', 'Recipe_Content__c'])
            writer.writeheader()
            writer.writerows(import_records)

        print()
        print("="*70)
        print("SALESFORCE IMPORT CSV CREATED!")
        print("="*70)
        print(f"File: {output_csv}")
        print()
        print(f"Total recipes in Excel: {ws.max_row - 1}")
        print(f"Matched to existing Salesforce meals: {matched_count}")
        print(f"New recipes (need Salesforce records): {new_records_count}")
        print(f"Skipped (no name): {skipped_count}")
        print(f"READY TO IMPORT: {len(import_records)} recipes")
        print()
        print("="*70)
        print("NEXT STEP:")
        print("="*70)
        print("1. Import via Salesforce Workbench:")
        print("   - Go to: https://workbench.developerforce.com")
        print("   - Data > Update")
        print("   - Object: Meal__c")
        print(f"   - Upload: {output_csv}")
        print("   - Map fields and update!")
        print()

        return import_records
    else:
        print("ERROR: No recipes matched to Salesforce IDs")
        return []

if __name__ == '__main__':
    reviewed_excel = r'c:\Users\Abbyl\OneDrive\Desktop\Salesforce Training\Assistant\Assistant\ALL_21_RECIPES_FIXED_Reviewed.xlsx'
    existing_meals_csv = r'c:\Users\Abbyl\OneDrive\Desktop\Salesforce Training\Assistant\Assistant\data\existing_meals_export.csv'
    output_csv = r'c:\Users\Abbyl\OneDrive\Desktop\Salesforce Training\Assistant\Assistant\data\Meal__c_FINAL_IMPORT.csv'

    print("="*70)
    print("CREATING SALESFORCE IMPORT CSV FROM REVIEWED FILE")
    print("="*70)
    print()

    records = create_salesforce_import_csv(reviewed_excel, existing_meals_csv, output_csv)

    print()
    print("="*70)
    print(f"SUCCESS! {len(records)} recipes ready for Salesforce import")
    print("="*70)
