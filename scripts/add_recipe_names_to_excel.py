"""
Add a new column to the Excel file showing the actual Salesforce meal names
This helps user verify that the recipe data matches the ID
"""
from openpyxl import load_workbook
import csv

def add_salesforce_names_column(user_excel, salesforce_csv, output_excel):
    """Add a column showing what Salesforce meal each ID corresponds to"""

    # Load Salesforce meals (ID → Name mapping)
    salesforce_meals = {}
    with open(salesforce_csv, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            salesforce_meals[row['Id'].strip()] = row['Name'].strip()

    print(f"Loaded {len(salesforce_meals)} Salesforce meals")

    # Load user's Excel file
    wb = load_workbook(user_excel)
    ws = wb.active

    # Add header in column B (currently blank)
    ws.cell(row=1, column=2, value="Salesforce_Meal_Name")

    # Fill in Salesforce meal names based on IDs in column A
    for row_num in range(2, ws.max_row + 1):
        sf_id = ws.cell(row=row_num, column=1).value

        if sf_id and isinstance(sf_id, str) and sf_id.startswith('a1'):
            if sf_id in salesforce_meals:
                sf_name = salesforce_meals[sf_id]
                ws.cell(row=row_num, column=2, value=sf_name)
                print(f"Row {row_num}: {sf_id} -> {sf_name}")
            else:
                ws.cell(row=row_num, column=2, value="[ID NOT FOUND]")
                print(f"Row {row_num}: {sf_id} -> NOT FOUND IN SALESFORCE")

    # Save the updated Excel file
    wb.save(output_excel)

    print()
    print("="*80)
    print("EXCEL FILE UPDATED!")
    print("="*80)
    print(f"Saved to: {output_excel}")
    print()
    print("Column B now shows the Salesforce meal name for each ID.")
    print("You can now review and verify that the recipe data matches the meal name.")
    print()

if __name__ == '__main__':
    user_excel = r'C:\Users\Abbyl\OneDrive\Desktop\Receips\new\existing_meals_export2.xlsx'
    salesforce_csv = r'C:\Users\Abbyl\OneDrive\Desktop\Salesforce Training\Assistant\Assistant\data\existing_meals_export.csv'
    output_excel = r'C:\Users\Abbyl\OneDrive\Desktop\Receips\new\existing_meals_export2_WITH_NAMES.xlsx'

    print("Adding Salesforce meal names to Excel file...")
    print()

    add_salesforce_names_column(user_excel, salesforce_csv, output_excel)

    print("="*80)
    print("NEXT STEP:")
    print("="*80)
    print(f"1. Open: {output_excel}")
    print("2. Review column B (Salesforce_Meal_Name)")
    print("3. Verify that the recipe data in columns C-E matches the meal name")
    print("4. If they DON'T match, you'll need to:")
    print("   - Either fix the ID in column A")
    print("   - Or remove that row from the import")
    print()
