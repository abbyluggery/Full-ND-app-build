"""
Verify that Salesforce IDs match the correct recipe names
Compare user's import file to actual Salesforce meal names
"""
from openpyxl import load_workbook
import csv
import difflib
import datetime

def verify_id_matches(user_excel, salesforce_csv, output_report):
    """Check if IDs in user's file match correct meal names in Salesforce"""

    # Load Salesforce meals (correct ID → Name mapping)
    salesforce_meals = {}
    with open(salesforce_csv, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            sf_id = row['Id'].strip()
            sf_name = row['Name'].strip()
            salesforce_meals[sf_id] = sf_name

    print(f"Loaded {len(salesforce_meals)} Salesforce meals")
    print()

    # Load user's Excel file
    wb = load_workbook(user_excel)
    ws = wb.active

    # Assume structure: Col A=Id, Col B=blank, Col C=Ingredients, Col D=Instructions, Col E=Recipe_Content
    # We need to figure out what recipe this is for

    mismatches = []
    matches = []
    unknown_ids = []

    print("="*80)
    print("VERIFICATION REPORT: Checking ID-to-Recipe Matches")
    print("="*80)
    print()

    for row_num in range(2, ws.max_row + 1):
        sf_id = ws.cell(row=row_num, column=1).value

        if not sf_id or not isinstance(sf_id, str) or not sf_id.startswith('a1'):
            continue

        # Get recipe data to try to identify what recipe this is
        ingredients = ws.cell(row=row_num, column=3).value or ''
        instructions = ws.cell(row=row_num, column=4).value or ''
        recipe_content = ws.cell(row=row_num, column=5).value or ''

        # Get the actual Salesforce meal name for this ID
        if sf_id in salesforce_meals:
            actual_sf_name = salesforce_meals[sf_id]
        else:
            unknown_ids.append({
                'row': row_num,
                'id': sf_id,
                'reason': 'ID not found in Salesforce'
            })
            print(f"[WARNING] Row {row_num}: UNKNOWN ID: {sf_id}")
            continue

        # Try to extract recipe name from the content
        # Look for patterns like "**Servings: 5 bowls**" or recipe names in content
        inferred_recipe_name = "Unknown"

        # Check if ingredients mention the recipe
        first_line = str(ingredients).split('\n')[0] if ingredients else ''
        if 'Ingredients' in first_line and '(' in first_line:
            # Pattern: "Ingredients (per bowl):" suggests this might be a bowl recipe
            if 'bowl' in first_line.lower():
                inferred_recipe_name = "Buddha Bowl or Bowl-type recipe"
            elif 'jar' in first_line.lower():
                inferred_recipe_name = "Jar recipe (Overnight Oats, Parfait, etc.)"
            elif 'wrap' in first_line.lower():
                inferred_recipe_name = "Wrap recipe"

        # Better: Look in Recipe_Content for the actual name
        if recipe_content and '**Servings:' in str(recipe_content):
            # Extract text before **Servings
            content_str = str(recipe_content)[:500]  # First 500 chars
            if 'Chicken' in content_str and 'Quinoa' in content_str:
                inferred_recipe_name = "Chicken & Quinoa Buddha Bowl"
            elif 'Overnight' in content_str and 'Oats' in content_str:
                inferred_recipe_name = "Overnight Oats (some variant)"
            elif 'Salmon' in content_str:
                inferred_recipe_name = "Salmon recipe"
            elif 'Lemon Pepper Chicken' in content_str:
                inferred_recipe_name = "Baked Lemon Pepper Chicken"
            elif 'Sheet Pan' in content_str and 'Chicken' in content_str:
                inferred_recipe_name = "Sheet Pan Chicken (variant)"
            elif 'Turkey' in content_str and 'Stir' in content_str:
                inferred_recipe_name = "Turkey Stir-Fry"
            elif 'Creamy' in content_str and 'Pasta' in content_str:
                inferred_recipe_name = "Creamy Pasta recipe"
            elif 'Fajitas' in content_str:
                inferred_recipe_name = "Chicken Fajitas"
            elif 'Soup' in content_str and 'Chicken' in content_str:
                inferred_recipe_name = "Chicken Soup"

        # Check if inferred name matches Salesforce name
        similarity = difflib.SequenceMatcher(None,
                                            inferred_recipe_name.lower(),
                                            actual_sf_name.lower()).ratio()

        result = {
            'row': row_num,
            'id': sf_id,
            'actual_sf_name': actual_sf_name,
            'inferred_name': inferred_recipe_name,
            'similarity': similarity,
            'ingredients_preview': str(ingredients)[:100] if ingredients else '',
            'content_preview': str(recipe_content)[:100] if recipe_content else ''
        }

        if similarity > 0.6:  # Likely a match
            matches.append(result)
            print(f"[OK] Row {row_num}: {sf_id} -> {actual_sf_name} (MATCH)")
        else:  # Potential mismatch
            mismatches.append(result)
            print(f"[MISMATCH] Row {row_num}: {sf_id} -> {actual_sf_name}")
            print(f"  Inferred recipe: {inferred_recipe_name}")
            print(f"  MISMATCH DETECTED!")
            print()

    # Generate report
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
    with open(output_report, 'w', encoding='utf-8') as f:
        f.write("# Recipe Import Verification Report\n\n")
        f.write(f"**Date:** {timestamp}\n\n")
        f.write("---\n\n")

        f.write("## Summary\n\n")
        f.write(f"- **Total Records:** {len(matches) + len(mismatches) + len(unknown_ids)}\n")
        f.write(f"- **Verified Matches:** {len(matches)}\n")
        f.write(f"- **MISMATCHES:** {len(mismatches)}\n")
        f.write(f"- **Unknown IDs:** {len(unknown_ids)}\n\n")
        f.write("---\n\n")

        if mismatches:
            f.write("## ⚠️ MISMATCHES DETECTED ⚠️\n\n")
            f.write("These records have IDs that don't match the recipe content:\n\n")
            for item in mismatches:
                f.write(f"### Row {item['row']}\n")
                f.write(f"- **Salesforce ID:** `{item['id']}`\n")
                f.write(f"- **Actual Salesforce Meal Name:** {item['actual_sf_name']}\n")
                f.write(f"- **Inferred Recipe from Content:** {item['inferred_name']}\n")
                f.write(f"- **Ingredients Preview:** {item['ingredients_preview']}...\n\n")

        if unknown_ids:
            f.write("## Unknown IDs\n\n")
            for item in unknown_ids:
                f.write(f"- Row {item['row']}: {item['id']} - {item['reason']}\n")
            f.write("\n")

        if matches:
            f.write("## ✓ Verified Matches\n\n")
            f.write(f"These {len(matches)} records appear to match correctly:\n\n")
            for item in matches[:10]:  # Show first 10
                f.write(f"- Row {item['row']}: {item['actual_sf_name']}\n")
            if len(matches) > 10:
                f.write(f"- ... and {len(matches) - 10} more\n\n")

    print()
    print("="*80)
    print("VERIFICATION COMPLETE")
    print("="*80)
    print(f"Report saved to: {output_report}")
    print()
    print(f"[OK] Verified matches: {len(matches)}")
    print(f"[ERROR] MISMATCHES: {len(mismatches)}")
    print(f"[WARNING] Unknown IDs: {len(unknown_ids)}")
    print()

    return {
        'matches': matches,
        'mismatches': mismatches,
        'unknown_ids': unknown_ids
    }

if __name__ == '__main__':
    user_excel = r'C:\Users\Abbyl\OneDrive\Desktop\Receips\new\existing_meals_export2.xlsx'
    salesforce_csv = r'C:\Users\Abbyl\OneDrive\Desktop\Salesforce Training\Assistant\Assistant\data\existing_meals_export.csv'
    output_report = r'C:\Users\Abbyl\OneDrive\Desktop\Salesforce Training\Assistant\Assistant\MISMATCH_REPORT.md'

    print("Verifying recipe ID matches...")
    print()

    results = verify_id_matches(user_excel, salesforce_csv, output_report)

    if results['mismatches']:
        print("="*80)
        print("ACTION REQUIRED - MISMATCHES DETECTED")
        print("="*80)
        print("Mismatches detected! Please review MISMATCH_REPORT.md")
        print("You'll need to fix these before importing.")
        print()
